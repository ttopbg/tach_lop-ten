import streamlit as st
import pandas as pd
import re
import io
import os
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Tách Lớp & Họ Tên", page_icon="🐍", layout="centered")

st.title("👽 Tách Lớp & Họ Tên")
st.markdown("Upload file Excel có sheet **Data**, cột **Họ và tên** dạng `12A01-Phạm Vũ Trường An`")


def remove_leading_zero(class_code):
    """12A01 → 12A1, 12B03 → 12B3, 12A10 → 12A10 (giữ nguyên nếu không có 0 dẫn đầu)"""
    if not isinstance(class_code, str):
        return class_code
    match = re.match(r'^(\d+)([A-Za-z]+)(\d+)$', class_code.strip())
    if match:
        prefix_num = match.group(1)
        letters = match.group(2)
        suffix_num = match.group(3)
        suffix_no_zero = str(int(suffix_num))
        return f"{prefix_num}{letters}{suffix_no_zero}"
    return class_code.strip()


def find_sheet_name(file_bytes):
    """Tìm sheet 'Data'/'data'/'DATA', nếu không có thì lấy sheet đầu tiên."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheets = wb.sheetnames
    wb.close()
    for s in sheets:
        if s.lower() == "data":
            return s, None  # (tên sheet, thông báo)
    # fallback: sheet đầu tiên
    return sheets[0], f"⚠️ Không tìm thấy sheet 'Data' — đang dùng sheet đầu tiên: **{sheets[0]}**"


def process_excel(file_bytes, filename):
    raw = file_bytes if isinstance(file_bytes, bytes) else file_bytes.read()

    sheet_name, warning = find_sheet_name(raw)
    if warning:
        st.warning(warning)

    try:
        wb_data = pd.read_excel(io.BytesIO(raw), sheet_name=sheet_name, dtype=str)
    except Exception as e:
        st.error(f"❌ Lỗi đọc file: {e}")
        return None, None

    # Tìm cột "Họ và tên"
    col_map = {c.strip(): c for c in wb_data.columns}
    target_col = col_map.get("Họ và tên")

    if target_col is None:
        st.error("❌ Không tìm thấy cột **'Họ và tên'** trong sheet Data!")
        return None, None

    col_idx = wb_data.columns.get_loc(target_col)

    def extract_name(val):
        if not isinstance(val, str) or '-' not in val:
            return val
        parts = val.split('-', 1)
        return parts[1].strip() if len(parts) == 2 else val

    def extract_class(val):
        if not isinstance(val, str) or '-' not in val:
            return val
        parts = val.split('-', 1)
        raw_class = parts[0].strip() if len(parts) == 2 else val
        return remove_leading_zero(raw_class)

    ho_ten_1 = wb_data[target_col].apply(extract_name)
    lop_1 = wb_data[target_col].apply(extract_class)

    wb_data.insert(col_idx + 1, "Họ tên HS", ho_ten_1)
    wb_data.insert(col_idx + 2, "Lớp CK", lop_1)

    base_name = os.path.splitext(filename)[0]
    out_name = f"{base_name}_đã tách lớp.xlsx"

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb_data.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        # Ẩn cột "Họ và tên" gốc (col_idx là 0-based, openpyxl dùng 1-based)
        ws.column_dimensions[get_column_letter(col_idx + 1)].hidden = True
        # Tự động giãn độ rộng cột theo nội dung, tối đa 50 ký tự
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    output.seek(0)

    return output, out_name


# ── UI ──────────────────────────────────────────────────────────────────────
uploaded_file = st.file_uploader("📂 Chọn file Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    st.info(f"📄 File đã chọn: **{uploaded_file.name}**")

    if st.button("⚙️ Xử lý", type="primary"):
        with st.spinner("Đang xử lý dữ liệu..."):
            result, out_name = process_excel(uploaded_file.read(), uploaded_file.name)

        if result:
            # Preview
            preview_df = pd.read_excel(result, dtype=str)
            result.seek(0)
            st.success(f"✅ Xử lý thành công! File output: **{out_name}**")

            # Show preview of new columns
            cols_to_show = ["Họ và tên", "Họ tên HS", "Lớp CK"]
            available = [c for c in cols_to_show if c in preview_df.columns]
            st.dataframe(preview_df[available].head(10), use_container_width=True)

            st.download_button(
                label="⬇️ Tải file đã xử lý",
                data=result,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

st.markdown("---")
# with st.expander("ℹ️ Hướng dẫn deploy lên Streamlit Cloud"):
#     st.markdown("""
# 1. Tạo repo GitHub, push 2 file: `app.py` và `requirements.txt`
# 2. Truy cập [share.streamlit.io](https://share.streamlit.io) → **New app**
# 3. Chọn repo, branch `main`, main file: `app.py` → **Deploy!**
#     """)
